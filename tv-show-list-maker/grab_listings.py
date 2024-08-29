#!/usr/bin/env python # [1]
import sys
import os
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors



def is_movie_file(path):
    valid_extensions = [".mkv", ".avi", ".mp4", ".webm", ".m4v"]
    if os.path.isfile(path):
        for ext in valid_extensions:
            if path.lower().endswith(ext):
                return True
    return False


def get_video_quality(height):
    if height < 460:
        return "SD"
    elif 460 <= height <= 480:
        return "480p"
    elif 480 < height <= 720:
        return "720p"
    elif 720 < height <= 1080:
        return "1080p"
    else:
        return "4K"


def process_show_folder(folder_path, show_name, current_row, sheet):

    potential_seasons = os.listdir(folder_path)

    for season in potential_seasons:
        season_path = os.path.join(folder_path, season)
        if os.path.isdir(season_path):
            episode_list = os.listdir(season_path)
            episode_count = 0
            for episode in episode_list:
                if is_movie_file(os.path.join(season_path, episode)):
                    episode_count += 1
            print(show_name + ":" + str(current_row))
            sheet.cell(row=current_row, column=1).value = show_name
            sheet.cell(row=current_row, column=2).value = season
            sheet.cell(row=current_row,
                       column=3).value = episode_count
            current_row += 1


def main():
    if len(sys.argv) < 2:
        print("Usage: python script_name.py directory_path")
        return

    location = os.path.abspath(sys.argv[1])

    currentRow = 2
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Television shows"

    # Headers
    sheet['A1'] = "Show"
    sheet['B1'] = "Season"
    sheet['C1'] = "Episode Count"
    sheet['D1'] = 'Notes'

    show_names = os.listdir(location)

    for season_name in show_names:
        season_path = os.path.join(location, season_name)
        if os.path.isdir(season_path):
            process_show_folder(season_path, season_name, currentRow, sheet)
            currentRow += 1

    output_file = 'show_list.xlsx'
    wb.save(output_file)
    print("Show list saved as", output_file)


if __name__ == "__main__":
    main()
