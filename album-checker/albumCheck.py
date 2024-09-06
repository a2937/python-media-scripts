#!/usr/bin/env python # [1]
import sys
import os
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors


def is_music_file(path):
    valid_extensions = [".mp3",".m4a",".flac"]
    if os.path.isfile(path):
        for ext in valid_extensions:
            if path.lower().endswith(ext):
                return True
    return False


def process_artist_albums(location, artist):
    Dictionary = {}
    potArtistPath = os.path.join(os.path.abspath(location), artist)
    print("Artist path: " + potArtistPath); 
    if os.path.isdir(potArtistPath):
        # Grab every album by that artist
        potentialAlbumNames = os.listdir(potArtistPath)
        for album in potentialAlbumNames:  # Look through every album
            potAlbumPath = os.path.join(os.path.abspath(
                  potArtistPath), album)  # Get the path of the album
            if os.path.isdir(potAlbumPath):
                songList = os.listdir(potAlbumPath)
                albumSongCount = 0
                Dictionary[artist + ";" + album] = 0
                for song in songList:
                    musicPath = os.path.join(os.path.abspath(
                          potAlbumPath), song)
                    if (is_music_file(musicPath)):
                        print(musicPath)
                        albumSongCount += 1

                Dictionary[artist + ";" + album] = albumSongCount
    return Dictionary


def main():
    if len(sys.argv) < 2:
        print("Usage: python script_name.py directory_path")
        return
    
    output_file = 'album_list.xlsx'
    location = os.path.abspath(sys.argv[1])
    yellowFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='FFFF00'))
    

    if(os.path.exists(output_file)):
        os.remove(output_file)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Albums"
  
    # Headers
    sheet['A1'] = "Artist"
    sheet['B1'] = "Album"
    sheet['C1'] = "Song"
    sheet['D1'] = "Status"
    sheet['E1'] = 'Notes'
    
    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["D"].width = 25

    artistNames = os.listdir(location)

    Dictionary = {}
    for artist in artistNames:  # loop through all the files and folders
      artist_album_data = process_artist_albums(location, artist)
      Dictionary.update(artist_album_data)
      
    for index, (artist, count) in enumerate(sorted(Dictionary.items(), key=lambda i: i[0].lower())): 
        currentRow = index + 2
        artistTrueName = artist.split(";")[0]
        albumName = artist.split(";")[1]
        sheet['A' + str(currentRow)] = artistTrueName
        sheet['B' + str(currentRow)] = albumName
        if int(count) > 3 : 
            sheet['B' + str(currentRow)].fill = yellowFill
            sheet['D' + str(currentRow)] = "Full"
        else:
            sheet['D' + str(currentRow)] = "Incomplete"
        sheet['C' + str(currentRow)] = count
        

    wb.save(output_file)
    print("Album list saved as", output_file)

if __name__ == "__main__":
    main()
