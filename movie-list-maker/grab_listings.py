#!/usr/bin/env python # [1]
import sys
import os
import cv2
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import colors


def is_movie_file(path):
    valid_extensions = [".mkv", ".avi", ".mp4", ".webm", ".m4v"]
    if os.path.isfile(path):
        for ext in valid_extensions:
            if path.lower().endswith(ext):
                return True
    return False


def get_video_quality(height):
    if height < 460:
        return "SD"
    elif 460 <= height <= 480:
        return "480p"
    elif 480 < height <= 720:
        return "720p"
    elif 720 < height <= 1080:
        return "1080p"
    else:
        return "4K"


def process_movie_folder(folder_path, folder_name):
    potential_file_names = os.listdir(folder_path)
    full_names = [os.path.join(folder_path, name)
                  for name in potential_file_names]
    movie_files = list(filter(is_movie_file, full_names))

    if movie_files:
        movie_path = movie_files[0]
        print(movie_path)
    else:
        print("No movie found in:", folder_name)
        return ["Unknown", 0, "Unknown"]

    try:
        vid = cv2.VideoCapture(movie_path)
        if vid.isOpened():
            height = vid.get(cv2.CAP_PROP_FRAME_HEIGHT)
            width = vid.get(cv2.CAP_PROP_FRAME_WIDTH)
            size = get_video_quality(height)
            vid.release()

            if height > 720 and width > 576:
                return ["Blu-ray", height, size]
            else:
                return ["DVD", height, size]
        else:
            print("Cannot read:", movie_path, "properly")
            return ["Unknown", 0, "Unknown"]
    except cv2.error as e:
        # Handle OpenCV-specific errors
        print("OpenCV error:", e)
        return ["Error", 0, "Unknown"]
    except Exception as e:
        # Handle other general exceptions
        print("An error occurred:", e)
        return ["Unknown", 0, "Unknown"]
    except:
        return ["Unknown", 0, "Unknown"]


def main():
    if len(sys.argv) < 2:
        print("Usage: python script_name.py directory_path")
        return

    output_file = 'movie_list.xlsx'
    location = os.path.abspath(sys.argv[1])

    if(os.path.exists(output_file)):
        os.remove(output_file)

    filenames = os.listdir(location)
    dictionary = {}
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Movies"

    # Headers
    sheet['A1'] = "Movie Name"
    sheet['B1'] = "Type"
    sheet['C1'] = "Height"
    sheet['D1'] = 'Resolution'
    sheet.column_dimensions["A"].width = 50


    try:
        for filename in filenames:
            potential_movie_path = os.path.join(location, filename)
            if os.path.isdir(potential_movie_path):
                size_info = process_movie_folder(
                    potential_movie_path, filename)
                dictionary[filename] = size_info

           

        # Save folder names and sizes to a file
        with open('fileListings.txt', 'w') as f:
            current_row = 2 
            for folder_name, size_info in sorted(dictionary.items()):
                f.write("%s : %s : %d : %s\n" %
                        (folder_name, size_info[0], size_info[1], size_info[2]))
                sheet.cell(row=current_row, column=1).value = folder_name
                sheet.cell(row=current_row, column=2).value = size_info[0]
                sheet.cell(row=current_row,
                      column=3).value = size_info[1]
                sheet.cell(row=current_row,
                      column=4).value = size_info[2]
                current_row = current_row + 1

        wb.save(output_file)
        print("Show list saved as", output_file)

      
    except Exception as e:
        print("An error occurred:", e)


if __name__ == "__main__":
    main()
